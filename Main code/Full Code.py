from pytube import YouTube
import os
import subprocess
import telebot
import time
from threading import Thread
from openpyxl import load_workbook


token = 'token'
bot = telebot.TeleBot(token)


class getSet:
	#------------Name-----------------
	def __init__(self):
		self.name = None

	def getName(self):
		return self.name

	def setName(self, name):
		self.name = name
	#-------------Path-----------------
	def __init__(self):
		self.path = None

	def getPath(self):
		return self.path

	def setPath(self, path):
		self.path = path
	#----------ID---------------------
	def __init__(self):
		self.user_id = None

	def getID(self):
		return self.user_id

	def setID(self, user_id):
		self.user_id = user_id
	#----------BoolTimer---------------------
	def __init__(self):
		self.boolTimer = None

	def getBoolTimer(self):
		return self.boolTimer

	def setBoolTimer(self, boolTimer):
		self.boolTimer = boolTimer

class primeSub:
	def addUpdateUser(self, user_id):
		self.userId = user_id.getID()
		self.wb = load_workbook("users.xlsx")
		self.sheet = self.wb.active
		lastRow = self.sheet.max_row

		self.listUsersId = [cellUser.value for cellUser in self.sheet["A"]]

		if self.userId in self.listUsersId:
			for cell in self.sheet["A"]:
				value = cell.value
				if self.userId == value:
					coordinates = cell.coordinate
					cellB = "B" + str(coordinates[-1])
					if self.sheet[cellB].value != 5:
						self.sheet[cellB].value += 1
						return True
					else:
						return False
		else: 
			cellA = "A" + str(lastRow + 1)
			cellB = "B" + str(lastRow + 1)
			self.sheet[cellA] = self.userId
			self.sheet[cellB] = 0
			self.sheet[cellB].value += 1
			return True

		self.wb.save("users.xlsx")

class timer:
	def start(self, message, waitMsg, name, boolTimer):
		seconds = 0
		self.name = name
		self.boolTimer = boolTimer
		while self.boolTimer.getBoolTimer() == True:
			time.sleep(1)
			seconds += 1
			bot.edit_message_text(chat_id=message.chat.id, message_id=waitMsg.message_id, text=f"Прошло {seconds} секунд. \n"
																							f"\n"
																							f"Скачиваю - {self.name.getName()}")
		bot.delete_message(chat_id=message.chat.id, message_id=waitMsg.message_id)

class startDownloadVideo:
	@bot.message_handler(commands=['download'])
	def Download(self, message, text, yt, name, boolTimer):
		self.path = getSet()

		self.path.setPath(yt.streams.filter(progressive=True, file_extension='mp4').order_by('resolution')[-1].download())

		self.audioFile = f"{name.getName()}.mp3"
		subprocess.call(["ffmpeg", "-i", self.path.getPath(), self.audioFile])

		bot.send_audio(message.chat.id, audio=open(self.audioFile, 'rb'), title=self.audioFile.replace(".mp3", ''))
		boolTimer.setBoolTimer(False)
		os.remove(self.audioFile)
		os.remove(self.path.getPath())


def error_message(message):
	bot.send_message(message.chat.id, 'Сервер не бесконечный. Не более 5 раз за день можешь скачивать\n')


@bot.message_handler(commands=['start'])
def start_message(message):
	bot.send_message(message.chat.id, 'Добро пожаловать!!!\n'
		'\n'
		'Чтобы скачать музыку с YouTube - отправьте ссылку на видео.\n')


@bot.message_handler()
def any_message(message):
	text = message.text
	correctLink1 = 'https://www.youtube.com/'
	correctLink2 = "https://youtu.be/"
	correctLin3 = "https://music.youtube.com/"

	if correctLink1 in text or correctLink2 in text or correctLin3 in text:
		try:
			yt = YouTube(text)
			duration = yt.length

			name = getSet()
			name.setName(yt.title)

			boolTimer = getSet()
			boolTimer.setBoolTimer(True)

			if duration < 600:
				user_id = getSet()
				user_id.setID(message.from_user.id)

				bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
				waitMsg = bot.send_message(message.chat.id, 'Пожалуйста, подождите...')

				checkPermission = primeSub().addUpdateUser(user_id)

				if checkPermission == True:
					threadOne = Thread(target=timer().start, args=(message, waitMsg, name, boolTimer))
					threadTwo = Thread(target=startDownloadVideo().Download, args=(message, text, yt, name, boolTimer))
					threadOne.start()
					threadTwo.start()
					return
				else:
					bot.delete_message(chat_id=message.chat.id, message_id=waitMsg.message_id)
					error_message(message)
					return
			else:
				bot.send_message(message.chat.id, text="Видео должно быть менее 10 минут")
		except:
			bot.send_message(message.chat.id, text="Некорректная ссылка")
	else:
	    bot.send_message(message.chat.id, text="Неправильная ссылка")




if __name__ == "__main__":
    print('We work')
    bot.polling()
